import calendar
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.styles.colors import Color

# Define the year
year = 2023

# Create a list of months
months = [
    "January", "February", "March", "April",
    "May", "June", "July", "August",
    "September", "October", "November", "December"
]

# Create a list of days of the week (first letter)
days_of_week = ["S", "M", "T", "W", "T", "F", "S"]

# Define the silver grey color (RGB: #C0C0C0)
silver_grey_color = Color(rgb="C0C0C0")

# Create a new workbook
workbook = openpyxl.Workbook()

# Select the active sheet
sheet = workbook.active

# Set the paper size to A3 and landscape orientation
sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

# Set the font and alignment for the year cell
year_font = Font(size=24, bold=True)
year_alignment = Alignment(horizontal="center")
year_header = sheet.cell(row=1, column=1, value=year)
year_header.font = year_font
year_header.alignment = year_alignment
year_header.fill = PatternFill(start_color=silver_grey_color, end_color=silver_grey_color, fill_type="solid")
sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(months))

# Set the border style for days and months
border_style = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

# Write the month row
for index, month in enumerate(months, start=1):
    month_cell = sheet.cell(row=3, column=index)
    month_cell.value = month

    # Set font size and center alignment
    month_cell.font = Font(size=16)
    month_cell.alignment = Alignment(horizontal="center")
    month_cell.border = border_style

    # Get the weekday of the first day of the month
    first_day = f"{year}-{index:02d}-01"
    first_day_weekday = calendar.weekday(year, index, 1)  # Monday is 0 and Sunday is 6

    # Determine the offset for the first day of the month
    offset = (first_day_weekday + 1) % 7

    # Get the number of days in the month
    num_days = calendar.monthrange(year, index)[1]

    # Write the day rows for each month
    for day in range(1, num_days + 1):
        day_cell = sheet.cell(row=3 + day, column=index)
        day_of_week = days_of_week[(offset + day - 1) % 7]
        day_cell.value = f"{day} {day_of_week}"
        day_cell.border = border_style

        # Check if the day is a weekend (Saturday)
        is_weekend = (day_of_week == "S")

        # Set the background color to silver grey for weekends
        if is_weekend:
            day_cell.fill = PatternFill(start_color=silver_grey_color, end_color=silver_grey_color, fill_type="solid")

# Save the workbook
workbook.save("calendar.xlsx")
